#!/usr/bin/python3

import logging, time, os
from openpyxl import load_workbook
from pathlib import Path


def print_all_state_data():
	logging.debug('Printing all states and counties.')
	time.sleep(0.003)
	for state, counties in census_data.items():
		print(state + ' state has the counties:')
		if debugging:
			print(counties)
		for county, county_data in counties.items():
			print('\t' + county + ' county has the data:')
			if debugging:
				print(county_data)
			for key, value in county_data.items():
				print('\t\tTotal ' + key + ': ' + str(value))


def print_state_data(state_name):
	logging.debug('Printing ' + state_name + ' and its counties if it exists.')
	time.sleep(0.003)
	if debugging:
		if state_name.upper() in census_data:
			print(state_name.upper() + ' state was found.')
			for state, counties in census_data.items():
				if state == state_name.upper():
					print(state + ' state has the counties:')
					for county, county_data in counties.items():
						print('\t' + county + ' county has the data:')
						for key, value in county_data.items():
							print('\t\tTotal ' + key + ': ' + str(value))

		else:
			print(state_name.upper() + ' state wasn''t found.')


def write_to_file(file_path, file_name):
	logging.info('Writing US Census statistic data for states and counties.')
	if not Path.exists(file_path):
		logging.info(str(file_path) + ' didn''t exist, so it was created.')
		Path.mkdir(file_path)
	# truncate if exists, create if not.
	file = open(file_path / file_name, 'w')
	time.sleep(0.003)
	for state, counties in census_data.items():
		file.write(state + ' state has the counties:\n')
		for county, county_data in counties.items():
			file.write('\t' + county + ' county has the data:\n')
			for key, value in county_data.items():
				file.write('\t\tTotal ' + key + ': ' + str(value) + '\n')


# Define logging
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - [%(levelname)s] - %(message)s')
debugging = True
if not debugging:
	logging.disable(logging.DEBUG)
logging.info('Starting UsCensusDataStatistics.py')
logging.info('Reading spreadsheet.')
# https://openpyxl.readthedocs.io/en/stable/tutorial.html#loading-from-a-file
workbook = load_workbook('../Examples/censuspopdata.xlsx')

logging.debug('Opening Worksheet.')
time.sleep(0.003)
# https://openpyxl.readthedocs.io/en/stable/usage.html#read-an-existing-workbook
worksheet = workbook.active

# Census tract is a geographic area defined for the Census.
census_data = {}
for row in range(2, worksheet.max_row + 1):
	state = worksheet['B' + str(row)].value
	county = worksheet['C' + str(row)].value
	population = worksheet['D' + str(row)].value

	# Create a state key if necessary, the key is the state abbreviation and the value is an empty dictionary
	census_data.setdefault(state, {})
	# Create a county key if necessary, the key is the county name and the value is a dictionary with 2 entries
	census_data[state].setdefault(county, {'tracts': 0, 'population': 0})
	# For each county row there is one census tract, keep a tally
	census_data[state][county]['tracts'] += 1
	# For each county row there is one population for each tract, keep a tally
	census_data[state][county]['population'] += int(population)

print_all_state_data()
print_state_data('az')
write_to_file(Path.cwd() / 'output', 'us-census-data.txt')
