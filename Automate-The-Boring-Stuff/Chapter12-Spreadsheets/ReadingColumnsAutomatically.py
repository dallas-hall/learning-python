#!/usr/bin/python3

import logging, sys, os, time

# Define logging output
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - [%(levelname)s] - %(message)s')

# Enable debugging messages
debugging = True
if not debugging:
	logging.disable(logging.DEBUG)
# Print start message and delay slightly
logging.info('Starting ' + os.path.relpath(sys.argv[0]))

logging.debug('Opening Workbook.')
# https://openpyxl.readthedocs.io/en/stable/tutorial.html#loading-from-a-file
from openpyxl import load_workbook
workbook = load_workbook('../Examples/example.xlsx')

logging.debug('Opening Worksheet.')
# https://openpyxl.readthedocs.io/en/stable/usage.html#read-an-existing-workbook
#worksheet = workbook.active
worksheet = workbook['Sheet1']

logging.debug('Printing Worksheet.')
time.sleep(.005)
for i in range(1, 8):
	print(i, worksheet.cell(row=i, column=2).value)

# Need plus 1 here otherwise the last row / column is missed.
for i in range(1, worksheet.max_row + 1):
	for j in range(1, worksheet.max_column + 1):
		print(str(i) + ',' + str(j), worksheet.cell(row=i, column=j).value)