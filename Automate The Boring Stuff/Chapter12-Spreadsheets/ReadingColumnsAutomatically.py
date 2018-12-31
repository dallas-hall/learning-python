#!/usr/bin/python3

# https://openpyxl.readthedocs.io/en/stable/tutorial.html#loading-from-a-file
from openpyxl import load_workbook
workbook = load_workbook('../Examples/example.xlsx')

# https://openpyxl.readthedocs.io/en/stable/usage.html#read-an-existing-workbook
worksheet = workbook.active

for i in range(1, 8):
	print(i, worksheet.cell(row=i, column=2).value)

# Need plus 1 here otherwise the last row / column is missed.
for i in range(1, worksheet.max_row + 1):
	for j in range(1, worksheet.max_column + 1):
		print(str(i) + ',' + str(j), worksheet.cell(row=i, column=j).value)