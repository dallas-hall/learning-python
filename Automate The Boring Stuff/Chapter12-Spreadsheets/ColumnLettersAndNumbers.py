#!/usr/bin/python3

import logging, sys, os, time

# Define logging output
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - [%(levelname)s] - %(message)s')

# Enable debugging messages
debugging = True
if not debugging:
	logging.disable(logging.DEBUG)
# Print start message and delay slightly
logging.info('Starting ' + os.path.relpath(sys.argv[0]))

# https://stackoverflow.com/a/36721735
from openpyxl.utils import get_column_letter, column_index_from_string

logging.debug('Printing generic column indices using openpyxl.utils')
time.sleep(.005)
print(get_column_letter(1))
print(get_column_letter(2))
print(column_index_from_string('A'))


logging.debug('Opening Workbook.')
time.sleep(.005)
# https://openpyxl.readthedocs.io/en/stable/optimized.html
from openpyxl import load_workbook
workbook = load_workbook('../Examples/example.xlsx', read_only=True)

logging.debug('Opening worksheet.')
time.sleep(.005)
# https://openpyxl.readthedocs.io/en/stable/usage.html#read-an-existing-workbook
worksheet = workbook['Sheet1']
logging.debug('Printing column indices from worksheet using openpyxl.utils')
# use the worksheet's maximum column to get the column letter and its number
print('Column number ' + str(column_index_from_string(get_column_letter(worksheet.max_column))) + ' has the column letter ' + get_column_letter(worksheet.max_column))





