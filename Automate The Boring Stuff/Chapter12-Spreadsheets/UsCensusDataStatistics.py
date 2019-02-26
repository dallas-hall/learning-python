#!/usr/bin/python3

import logging, time
from openpyxl import load_workbook


def print_all_counties():
	logging.debug('Printing all counties.')
	time.sleep(0.003)
	for state, counties in census_data.items():
		print(state + ' state has the counties:')
		if debugging:
			print(counties)
		for county, county_data in counties.items():
			print(county + ' county has the data:')
			if debugging:
				print(county_data)
			for key, value in county_data.items():
				print(key + ': ' + str(value))




# Define logging
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - [%(levelname)s] - %(message)s')
debugging = True
if not debugging:
	logging.disable(logging.DEBUG)
logging.info('Starting UsCensusDataStatistics.py')
logging.info('Reading spreadsheet.')
# https://openpyxl.readthedocs.io/en/stable/tutorial.html#loading-from-a-file
workbook = load_workbook('../Examples/censuspopdata.xlsx')

logging.debug('Opening Worksheet.')
time.sleep(0.003)
# https://openpyxl.readthedocs.io/en/stable/usage.html#read-an-existing-workbook
worksheet = workbook.active

# Census tract is a geographic area defined for the Census.
census_data = {}
for row in range(2, worksheet.max_row + 1):
	state = worksheet['B' + str(row)].value
	county = worksheet['C' + str(row)].value
	population = worksheet['D' + str(row)].value

	# Create a state key if necessary
	census_data.setdefault(state, {})
	# Create a county key if necessary
	census_data[state].setdefault(county, {'tracts': 0, 'population': 0})
	# Each row is one census tract
	census_data[state][county]['tracts'] += 1
	# Update county population
	census_data[state][county]['population'] += int(population)

print_all_counties()